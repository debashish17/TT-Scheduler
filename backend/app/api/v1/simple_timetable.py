"""
Simple timetable generation API.
Single endpoint that accepts all problem data and returns a complete timetable.
No pre-existing database records required.
"""
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from typing import Dict, List, Any, Optional
import logging

from app.core.simple_solver import solve_timetable

logger = logging.getLogger(__name__)
router = APIRouter()


class SubjectIn(BaseModel):
    name: str
    code: str
    periods_per_week: int = Field(default=3, ge=1, le=15)
    target_classes: List[str] = []


class TeacherIn(BaseModel):
    name: str
    subjects: List[str] = []


class ClassIn(BaseModel):
    name: str
    size: int = Field(default=30, ge=1, le=500)


class RoomIn(BaseModel):
    name: str
    capacity: int = Field(default=40, ge=1, le=1000)


class ConstraintsIn(BaseModel):
    max_consecutive_periods: int = Field(default=3, ge=1, le=10)
    lunch_after_period: int = Field(default=0, ge=0)
    max_periods_per_day_per_teacher: int = Field(default=8, ge=1, le=15)


class SimpleTimetableRequest(BaseModel):
    institution_name: str = "My School"
    subjects: List[SubjectIn]
    teachers: List[TeacherIn] = []
    classes: List[ClassIn]
    rooms: List[RoomIn]
    working_days: List[str] = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    periods_per_day: int = Field(default=7, ge=1, le=15)
    period_duration_minutes: int = Field(default=45, ge=15, le=180)
    start_time: str = "08:00"
    constraints: ConstraintsIn = Field(default_factory=ConstraintsIn)


@router.post("/generate-simple")
@router.post("/generate")
async def generate_simple_timetable(request: SimpleTimetableRequest):
    """
    Generate a school timetable from self-contained request data.
    The CP-SAT solver ensures no conflicts and spreads subjects across the week.
    """
    if not request.subjects:
        raise HTTPException(status_code=400, detail="At least one subject is required")
    if not request.classes:
        raise HTTPException(status_code=400, detail="At least one class is required")
    if not request.rooms:
        raise HTTPException(status_code=400, detail="At least one room is required")
    if not request.working_days:
        raise HTTPException(status_code=400, detail="Working days must be specified")

    teachers_data = request.teachers
    if not teachers_data:
        teachers_data = [
            TeacherIn(name=f"Teacher ({s.name})", subjects=[s.code])
            for s in request.subjects
        ]

    total_required = sum(s.periods_per_week for s in request.subjects) * len(request.classes)
    total_available = len(request.working_days) * request.periods_per_day * len(request.rooms)
    if total_required > total_available:
        logger.warning(f"Tight schedule: {total_required} sessions needed, {total_available} room-periods available")

    problem = {
        "institution_name": request.institution_name,
        "subjects": [s.model_dump() for s in request.subjects],
        "teachers": [t.model_dump() for t in teachers_data],
        "classes": [c.model_dump() for c in request.classes],
        "rooms": [r.model_dump() for r in request.rooms],
        "working_days": request.working_days,
        "periods_per_day": request.periods_per_day,
        "period_duration_minutes": request.period_duration_minutes,
        "start_time": request.start_time,
        "constraints": request.constraints.model_dump()
    }

    logger.info(
        f"Generating timetable for '{request.institution_name}': "
        f"{len(request.subjects)} subjects, {len(teachers_data)} teachers, "
        f"{len(request.classes)} classes, {len(request.rooms)} rooms"
    )

    try:
        result = solve_timetable(problem)
    except Exception as e:
        logger.error(f"Timetable generation error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Solver error: {str(e)}")

    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "Timetable generation failed"))

    return result


# ─────────────────────────────────────────────────────────────
# ANALYTICS  (computed directly from solver result — no DB)
# ─────────────────────────────────────────────────────────────

class AnalyticsRequest(BaseModel):
    assignments: List[Dict[str, Any]] = []
    working_days: List[str] = []
    time_slots: List[Dict[str, Any]] = []
    stats: Dict[str, Any] = {}
    solver: str = "CP-SAT"
    status: str = "FEASIBLE"
    solve_time: float = 0.0


@router.post("/analytics")
async def compute_timetable_analytics(body: AnalyticsRequest):
    """
    Compute analytics for a generated timetable.
    Returns teacher workloads, subject distribution, room usage, day load.
    """
    assignments = body.assignments
    working_days = body.working_days
    time_slots = body.time_slots

    if not assignments:
        return {
            "summary": {"total_assignments": 0},
            "teacher_workload": [],
            "subject_distribution": [],
            "room_usage": [],
            "day_load": [],
        }

    unique_classes = list({a["class_name"] for a in assignments})
    unique_teachers = list({a["teacher_name"] for a in assignments})
    unique_subjects = list({a["subject_code"] for a in assignments})
    unique_rooms = list({a["room_name"] for a in assignments})

    summary = {
        "total_assignments": len(assignments),
        "classes_count": len(unique_classes),
        "teachers_count": len(unique_teachers),
        "subjects_count": len(unique_subjects),
        "rooms_count": len(unique_rooms),
        "working_days_count": len(working_days),
        "periods_per_day": len(time_slots),
        "solver": body.solver,
        "solver_status": body.status,
        "solve_time_seconds": body.solve_time,
    }

    # Teacher workload
    teacher_map: Dict[str, Any] = {}
    for a in assignments:
        t = a["teacher_name"]
        if t not in teacher_map:
            teacher_map[t] = {"teacher_name": t, "periods_per_week": 0, "subjects": set(), "classes": set()}
        teacher_map[t]["periods_per_week"] += 1
        teacher_map[t]["subjects"].add(a["subject_code"])
        teacher_map[t]["classes"].add(a["class_name"])

    total_slots_per_teacher = len(working_days) * len(time_slots) if time_slots else 1
    teacher_workload = []
    for t, d in sorted(teacher_map.items()):
        ppw = d["periods_per_week"]
        teacher_workload.append({
            "teacher_name": t,
            "periods_per_week": ppw,
            "subjects": sorted(d["subjects"]),
            "classes": sorted(d["classes"]),
            "utilization_pct": round(ppw / max(total_slots_per_teacher, 1) * 100, 1),
            "overloaded": ppw > total_slots_per_teacher * 0.8,
        })

    # Subject distribution
    subj_map: Dict[str, Any] = {}
    for a in assignments:
        key = a["subject_code"]
        if key not in subj_map:
            subj_map[key] = {
                "subject_code": key,
                "subject_name": a.get("subject_name", key),
                "total_periods": 0,
                "classes_covered": set(),
            }
        subj_map[key]["total_periods"] += 1
        subj_map[key]["classes_covered"].add(a["class_name"])

    subject_distribution = [
        {
            "subject_code": code,
            "subject_name": d["subject_name"],
            "total_periods": d["total_periods"],
            "classes_covered": len(d["classes_covered"]),
            "avg_periods_per_class": round(d["total_periods"] / max(len(d["classes_covered"]), 1), 1),
        }
        for code, d in sorted(subj_map.items())
    ]

    # Room usage
    total_week_slots = len(working_days) * len(time_slots) if time_slots else 1
    room_map: Dict[str, int] = {}
    for a in assignments:
        r = a["room_name"]
        room_map[r] = room_map.get(r, 0) + 1

    room_usage = [
        {
            "room_name": r,
            "assignments": cnt,
            "utilization_pct": round(cnt / total_week_slots * 100, 1),
        }
        for r, cnt in sorted(room_map.items(), key=lambda x: -x[1])
    ]

    # Day load
    day_map: Dict[str, int] = {}
    for a in assignments:
        d = a.get("day", "Unknown")
        day_map[d] = day_map.get(d, 0) + 1

    day_load = [
        {"day": day, "sessions": day_map.get(day, 0)}
        for day in (working_days if working_days else sorted(day_map.keys()))
    ]

    return {
        "summary": summary,
        "teacher_workload": teacher_workload,
        "subject_distribution": subject_distribution,
        "room_usage": room_usage,
        "day_load": day_load,
    }


# ─────────────────────────────────────────────────────────────
# EXPORT  (returns .xlsx binary — no DB required)
# ─────────────────────────────────────────────────────────────

class ExportRequest(BaseModel):
    institution_name: str = "My School"
    assignments: List[Dict[str, Any]] = []
    working_days: List[str] = []
    time_slots: List[Dict[str, Any]] = []
    stats: Dict[str, Any] = {}


@router.post("/export/excel")
async def export_timetable_excel(body: ExportRequest):
    """
    Export the generated timetable as an Excel workbook.
    Returns an .xlsx with one sheet per class + a Faculty sheet + Summary sheet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        from fastapi.responses import StreamingResponse
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    assignments = body.assignments
    working_days = body.working_days
    time_slots = body.time_slots
    institution_name = body.institution_name

    if not assignments:
        raise HTTPException(status_code=400, detail="No assignments to export")

    # Build lookups
    class_lookup: Dict[tuple, Dict] = {}
    teacher_lookup: Dict[tuple, Dict] = {}
    for a in assignments:
        class_lookup[(a["class_name"], a["day"], a["period"])] = a
        teacher_lookup[(a["teacher_name"], a["day"], a["period"])] = a

    unique_classes = sorted({a["class_name"] for a in assignments})
    unique_teachers = sorted({a["teacher_name"] for a in assignments})
    unique_subjects = sorted({a["subject_code"] for a in assignments})

    SUBJECT_COLORS = [
        "DBEAFE", "D1FAE5", "EDE9FE", "FEF3C7", "FCE7F3",
        "FEF9C3", "CFFAFE", "F3E8FF", "ECFCCB", "FFEDD5",
    ]
    subj_color_map = {code: SUBJECT_COLORS[i % len(SUBJECT_COLORS)] for i, code in enumerate(unique_subjects)}

    NAVY = "1E3A5F"
    STEEL = "2563EB"
    WHITE = "FFFFFF"

    def hdr_cell(ws, row, col, val, bg=NAVY, fg=WHITE, bold=True, sz=10, wrap=False):
        c = ws.cell(row, col, val)
        c.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = thin_border()
        return c

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def data_cell(ws, row, col, val, bg="FFFFFF", wrap=True):
        c = ws.cell(row, col, val)
        c.font = Font(name="Calibri", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = thin_border()
        return c

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ──────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.merge_cells("A1:D1")
    t = ws_s["A1"]
    t.value = f"{institution_name} — Timetable Report"
    t.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_s.row_dimensions[1].height = 32
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 20

    stats = body.stats or {}
    rows = [
        ("Total Assignments", len(assignments)),
        ("Classes", len(unique_classes)),
        ("Teachers", len(unique_teachers)),
        ("Subjects", len(unique_subjects)),
        ("Rooms", len({a["room_name"] for a in assignments})),
        ("Working Days", len(working_days)),
        ("Periods per Day", len(time_slots)),
        ("Solver", stats.get("solver", "CP-SAT")),
        ("Solve Time (s)", stats.get("solve_time_seconds", "")),
    ]
    for i, (lbl, val) in enumerate(rows, start=2):
        ws_s.cell(i, 1, lbl).font = Font(name="Calibri", bold=True, size=10)
        ws_s.cell(i, 2, str(val)).font = Font(name="Calibri", size=10)

    # ── Class sheets ───────────────────────────────────────────
    for cls in unique_classes:
        safe = (cls[:28].replace("/", "-").replace("\\", "-")
                .replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", ""))
        ws = wb.create_sheet(safe)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(working_days) + 1)
        t = ws.cell(1, 1, f"{institution_name}  |  {cls}")
        t.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Column headers
        ws.column_dimensions["A"].width = 16
        hdr_cell(ws, 2, 1, "Period", sz=10)
        for ci, day in enumerate(working_days, start=2):
            hdr_cell(ws, 2, ci, day, sz=10)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

        # Period rows
        for ri, slot in enumerate(time_slots, start=3):
            pnum = slot.get("period", ri - 2)
            lbl = f"P{pnum}\n{slot.get('start','')}–{slot.get('end','')}"
            hdr_cell(ws, ri, 1, lbl, bg="334155", sz=9, wrap=True)
            ws.row_dimensions[ri].height = 44

            for ci, day in enumerate(working_days, start=2):
                a = class_lookup.get((cls, day, pnum))
                if a:
                    val = f"{a['subject_code']}\n{a.get('subject_name', '')}\n{a['teacher_name']}\n{a['room_name']}"
                    data_cell(ws, ri, ci, val, bg=subj_color_map.get(a["subject_code"], "FFFFFF"))
                else:
                    data_cell(ws, ri, ci, "")

    # ── Faculty sheet ──────────────────────────────────────────
    ws_f = wb.create_sheet("Faculty")
    ws_f.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(working_days) + 1)
    t = ws_f.cell(1, 1, f"{institution_name}  |  Faculty Schedule")
    t.font = Font(name="Calibri", bold=True, size=12, color=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_f.row_dimensions[1].height = 28
    ws_f.column_dimensions["A"].width = 22
    for ci, day in enumerate(working_days, start=2):
        hdr_cell(ws_f, 2, ci, day, sz=10)
        ws_f.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    cur_row = 3
    for teacher in unique_teachers:
        # Teacher name row
        ws_f.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(working_days) + 1)
        tc = ws_f.cell(cur_row, 1, teacher)
        tc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        tc.fill = PatternFill("solid", fgColor=STEEL)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws_f.row_dimensions[cur_row].height = 20
        cur_row += 1

        for slot in time_slots:
            pnum = slot.get("period", 1)
            lbl = f"P{pnum} {slot.get('start','')}–{slot.get('end','')}"
            hdr_cell(ws_f, cur_row, 1, lbl, bg="334155", sz=9)
            ws_f.row_dimensions[cur_row].height = 40
            for ci, day in enumerate(working_days, start=2):
                a = teacher_lookup.get((teacher, day, pnum))
                if a:
                    val = f"{a['subject_code']}\n{a['class_name']}\n{a['room_name']}"
                    data_cell(ws_f, cur_row, ci, val, bg=subj_color_map.get(a["subject_code"], "FFFFFF"))
                else:
                    data_cell(ws_f, cur_row, ci, "")
            cur_row += 1
        cur_row += 1  # gap between teachers

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_inst = institution_name.replace(" ", "_").replace("/", "-")
    filename = f"{safe_inst}_Timetable.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────
# SAVE TO DATABASE
# Persists a generated timetable to Supabase (timetables + timetable_entries)
# ─────────────────────────────────────────────────────────────

class SaveTimetableRequest(BaseModel):
    institution_name: str = "My School"
    name: str = "Timetable"
    solver: str = "CP-SAT"
    status: str = "FEASIBLE"
    solve_time: float = 0.0
    assignments: List[Dict[str, Any]] = []
    working_days: List[str] = []
    periods_per_day: int = 7
    stats: Dict[str, Any] = {}


@router.post("/save")
async def save_timetable_to_db(body: SaveTimetableRequest):
    """
    Persist a generated timetable to the Supabase database.
    Creates one row in `timetables` and one row per assignment in `timetable_entries`.
    Returns the saved timetable ID.
    """
    import uuid
    from datetime import datetime as dt

    try:
        from app.db.session import SessionLocal
        from sqlalchemy import text

        db = SessionLocal()
        try:
            timetable_id = str(uuid.uuid4())
            now = dt.utcnow().isoformat()

            # ── Insert into timetables ──────────────────────────
            db.execute(
                text("""
                    INSERT INTO timetables (
                        id, institution_id, name, status,
                        generation_params, metrics, created_at, updated_at
                    ) VALUES (
                        :id, NULL, :name, :status,
                        :params::jsonb, :metrics::jsonb, :created_at, :updated_at
                    )
                """),
                {
                    "id":         timetable_id,
                    "name":       body.name or f"{body.institution_name} Timetable",
                    "status":     "published",
                    "params":     __import__("json").dumps({
                        "institution_name": body.institution_name,
                        "working_days":     body.working_days,
                        "periods_per_day":  body.periods_per_day,
                        "solver":           body.solver,
                        "solver_status":    body.status,
                    }),
                    "metrics":    __import__("json").dumps(body.stats),
                    "created_at": now,
                    "updated_at": now,
                }
            )

            # ── Insert timetable_entries ────────────────────────
            entries_inserted = 0
            for a in body.assignments:
                entry_id = str(uuid.uuid4())
                db.execute(
                    text("""
                        INSERT INTO timetable_entries (
                            id, timetable_id, day_of_week, start_time, end_time,
                            period_number, entry_type, metadata, created_at, updated_at
                        ) VALUES (
                            :id, :timetable_id, :day, :start_time, :end_time,
                            :period, 'regular', :metadata::jsonb, :created_at, :updated_at
                        )
                    """),
                    {
                        "id":           entry_id,
                        "timetable_id": timetable_id,
                        "day":          a.get("day", ""),
                        "start_time":   a.get("start_time", ""),
                        "end_time":     a.get("end_time", ""),
                        "period":       a.get("period", 1),
                        "metadata":     __import__("json").dumps({
                            "class_name":   a.get("class_name"),
                            "subject_name": a.get("subject_name"),
                            "subject_code": a.get("subject_code"),
                            "teacher_name": a.get("teacher_name"),
                            "room_name":    a.get("room_name"),
                        }),
                        "created_at":   now,
                        "updated_at":   now,
                    }
                )
                entries_inserted += 1

            db.commit()
            logger.info(
                f"Timetable saved: id={timetable_id}, "
                f"{entries_inserted} entries written to DB"
            )

            return {
                "success":          True,
                "timetable_id":     timetable_id,
                "entries_saved":    entries_inserted,
                "message":          f"Timetable saved with {entries_inserted} entries",
            }

        except Exception as db_err:
            db.rollback()
            logger.error(f"DB save error: {db_err}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Database save failed: {str(db_err)}"
            )
        finally:
            db.close()

    except ImportError as e:
        raise HTTPException(status_code=503, detail=f"DB not available: {e}")

