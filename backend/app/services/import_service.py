"""
Import service for bulk data operations.
Handles Excel import for Faculty, Courses, Rooms, and other entities.
"""
import io
from typing import Dict, List, Any, Optional
from uuid import UUID
from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException

from app.utils.excel import ExcelReader, ExcelValidator, ExcelTemplateGenerator, ImportResult
from app.models import Institution, Department, Faculty, Course, Classroom, StudentBatch
from app.schemas.faculty import FacultyCreate
from app.schemas.course import CourseCreate
from app.schemas.room import RoomCreate
from app.schemas.batch import StudentBatchCreate
from app.services.base_service import BaseService


class ImportService:
    """Service for handling bulk data import operations."""

    def __init__(self):
        self.excel_reader = ExcelReader()
        self.template_generator = ExcelTemplateGenerator()

    async def import_faculty_excel(
        self,
        db: Session,
        file: UploadFile,
        institution_id: UUID
    ) -> ImportResult:
        """
        Import faculty from Excel file.

        Expected columns:
        - employee_id (required)
        - name (required)
        - email (optional)
        - department_code (required)
        - designation (optional)
        - max_hours_per_week (optional, default 18)
        - subjects_can_teach (optional, comma-separated)
        """
        # Define expected columns
        required_columns = ['employee_id', 'name', 'department_code']
        optional_columns = ['email', 'designation', 'max_hours_per_week', 'subjects_can_teach']

        validator = ExcelValidator(required_columns, optional_columns)
        result = ImportResult()

        try:
            # Validate file
            validator.validate_file_format(file)

            # Read Excel data
            headers, rows_data = await self.excel_reader.read_excel_file(file)

            # Validate headers
            missing_columns = validator.validate_headers(headers)
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns: {', '.join(missing_columns)}"
                )

            # Get institution and departments for validation
            institution = db.query(Institution).filter(Institution.id == institution_id).first()
            if not institution:
                raise HTTPException(status_code=404, detail="Institution not found")

            departments = db.query(Department).filter(
                Department.institution_id == institution_id,
                Department.deleted_at.is_(None)
            ).all()
            dept_code_to_id = {dept.code.upper(): dept.id for dept in departments}

            # Process each row
            for row_number, row_data in enumerate(rows_data, start=2):  # Start from 2 (Excel row number)
                try:
                    # Normalize data
                    clean_data = validator.normalize_row_data(row_data)

                    # Validate required fields
                    if not clean_data.get('employee_id'):
                        result.add_error(row_number, "Employee ID is required", clean_data)
                        continue

                    if not clean_data.get('name'):
                        result.add_error(row_number, "Name is required", clean_data)
                        continue

                    if not clean_data.get('department_code'):
                        result.add_error(row_number, "Department code is required", clean_data)
                        continue

                    # Validate department exists
                    dept_code = str(clean_data['department_code']).upper()
                    if dept_code not in dept_code_to_id:
                        result.add_error(
                            row_number,
                            f"Department '{dept_code}' not found in institution",
                            clean_data
                        )
                        continue

                    # Check for duplicate employee_id
                    existing_faculty = db.query(Faculty).filter(
                        Faculty.employee_id == clean_data['employee_id'],
                        Faculty.institution_id == institution_id,
                        Faculty.deleted_at.is_(None)
                    ).first()

                    if existing_faculty:
                        result.add_error(
                            row_number,
                            f"Faculty with employee ID '{clean_data['employee_id']}' already exists",
                            clean_data
                        )
                        continue

                    # Prepare faculty data
                    faculty_data = {
                        'institution_id': institution_id,
                        'department_id': dept_code_to_id[dept_code],
                        'employee_id': clean_data['employee_id'],
                        'name': clean_data['name'],
                        'email': clean_data.get('email'),
                        'designation': clean_data.get('designation'),
                        'max_hours_per_week': int(clean_data.get('max_hours_per_week', 18)),
                        'subjects_can_teach': []
                    }

                    # Process subjects_can_teach
                    if clean_data.get('subjects_can_teach'):
                        subjects_str = str(clean_data['subjects_can_teach'])
                        faculty_data['subjects_can_teach'] = [
                            s.strip() for s in subjects_str.split(',') if s.strip()
                        ]

                    # Validate with Pydantic schema
                    faculty_schema = FacultyCreate(**faculty_data)

                    # Create faculty record
                    faculty = Faculty(**faculty_schema.model_dump())
                    db.add(faculty)
                    db.flush()  # Get ID without committing

                    result.add_success({
                        'employee_id': faculty.employee_id,
                        'name': faculty.name,
                        'department_code': dept_code,
                        'id': str(faculty.id)
                    })

                except ValueError as e:
                    result.add_error(row_number, f"Data validation error: {str(e)}", clean_data)
                except Exception as e:
                    result.add_error(row_number, f"Unexpected error: {str(e)}", clean_data)

            # Commit if we have successful records
            if result.successful > 0:
                db.commit()
            else:
                db.rollback()

        except HTTPException:
            db.rollback()
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

        return result

    async def import_courses_excel(
        self,
        db: Session,
        file: UploadFile,
        institution_id: UUID
    ) -> ImportResult:
        """
        Import courses from Excel file.

        Expected columns:
        - code (required)
        - name (required)
        - department_code (required)
        - course_type (optional, default 'theory')
        - theory_credits (optional)
        - lab_credits (optional)
        - hours_per_week (required)
        - faculty_employee_id (optional)
        - batch_name (optional)
        - expected_students (optional)
        - required_features (optional, comma-separated)
        """
        required_columns = ['code', 'name', 'department_code', 'hours_per_week']
        optional_columns = [
            'course_type', 'theory_credits', 'lab_credits', 'faculty_employee_id',
            'batch_name', 'expected_students', 'required_features'
        ]

        validator = ExcelValidator(required_columns, optional_columns)
        result = ImportResult()

        try:
            validator.validate_file_format(file)
            headers, rows_data = await self.excel_reader.read_excel_file(file)

            missing_columns = validator.validate_headers(headers)
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns: {', '.join(missing_columns)}"
                )

            # Get reference data
            departments = db.query(Department).filter(
                Department.institution_id == institution_id,
                Department.deleted_at.is_(None)
            ).all()
            dept_code_to_id = {dept.code.upper(): dept.id for dept in departments}

            faculty = db.query(Faculty).filter(
                Faculty.institution_id == institution_id,
                Faculty.deleted_at.is_(None)
            ).all()
            faculty_emp_to_id = {f.employee_id.upper(): f.id for f in faculty}

            batches = db.query(StudentBatch).filter(
                StudentBatch.institution_id == institution_id,
                StudentBatch.deleted_at.is_(None)
            ).all()
            batch_name_to_id = {b.batch_name.upper(): b.id for b in batches}

            # Process rows
            for row_number, row_data in enumerate(rows_data, start=2):
                try:
                    clean_data = validator.normalize_row_data(row_data)

                    # Validate required fields
                    if not clean_data.get('code'):
                        result.add_error(row_number, "Course code is required", clean_data)
                        continue

                    if not clean_data.get('name'):
                        result.add_error(row_number, "Course name is required", clean_data)
                        continue

                    if not clean_data.get('department_code'):
                        result.add_error(row_number, "Department code is required", clean_data)
                        continue

                    if not clean_data.get('hours_per_week'):
                        result.add_error(row_number, "Hours per week is required", clean_data)
                        continue

                    # Validate department exists
                    dept_code = str(clean_data['department_code']).upper()
                    if dept_code not in dept_code_to_id:
                        result.add_error(
                            row_number,
                            f"Department '{dept_code}' not found",
                            clean_data
                        )
                        continue

                    # Check for duplicate course code
                    existing_course = db.query(Course).filter(
                        Course.code == clean_data['code'],
                        Course.institution_id == institution_id,
                        Course.deleted_at.is_(None)
                    ).first()

                    if existing_course:
                        result.add_error(
                            row_number,
                            f"Course with code '{clean_data['code']}' already exists",
                            clean_data
                        )
                        continue

                    # Prepare course data
                    course_data = {
                        'institution_id': institution_id,
                        'department_id': dept_code_to_id[dept_code],
                        'code': clean_data['code'],
                        'name': clean_data['name'],
                        'course_type': clean_data.get('course_type', 'theory'),
                        'theory_credits': float(clean_data.get('theory_credits', 0)),
                        'lab_credits': float(clean_data.get('lab_credits', 0)),
                        'hours_per_week': int(clean_data['hours_per_week']),
                        'expected_students': int(clean_data.get('expected_students', 0)),
                        'required_features': []
                    }

                    # Handle optional faculty assignment
                    if clean_data.get('faculty_employee_id'):
                        faculty_emp_id = str(clean_data['faculty_employee_id']).upper()
                        if faculty_emp_id in faculty_emp_to_id:
                            course_data['assigned_faculty_id'] = faculty_emp_to_id[faculty_emp_id]
                        else:
                            result.add_warning(
                                row_number,
                                f"Faculty '{faculty_emp_id}' not found, course created without faculty",
                                clean_data
                            )

                    # Handle optional batch assignment
                    if clean_data.get('batch_name'):
                        batch_name = str(clean_data['batch_name']).upper()
                        if batch_name in batch_name_to_id:
                            course_data['assigned_batch_id'] = batch_name_to_id[batch_name]
                        else:
                            result.add_warning(
                                row_number,
                                f"Batch '{batch_name}' not found, course created without batch",
                                clean_data
                            )

                    # Process required features
                    if clean_data.get('required_features'):
                        features_str = str(clean_data['required_features'])
                        course_data['required_features'] = [
                            f.strip().lower() for f in features_str.split(',') if f.strip()
                        ]

                    # Validate with Pydantic schema
                    course_schema = CourseCreate(**course_data)

                    # Create course record
                    course = Course(**course_schema.model_dump())
                    db.add(course)
                    db.flush()

                    result.add_success({
                        'code': course.code,
                        'name': course.name,
                        'department_code': dept_code,
                        'id': str(course.id)
                    })

                except ValueError as e:
                    result.add_error(row_number, f"Data validation error: {str(e)}", clean_data)
                except Exception as e:
                    result.add_error(row_number, f"Unexpected error: {str(e)}", clean_data)

            # Commit or rollback
            if result.successful > 0:
                db.commit()
            else:
                db.rollback()

        except HTTPException:
            db.rollback()
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

        return result

    async def import_rooms_excel(
        self,
        db: Session,
        file: UploadFile,
        institution_id: UUID
    ) -> ImportResult:
        """
        Import rooms from Excel file.

        Expected columns:
        - room_number (required)
        - building (optional)
        - capacity (required)
        - room_type (optional, default 'lecture_hall')
        - features (optional, comma-separated)
        """
        required_columns = ['room_number', 'capacity']
        optional_columns = ['building', 'room_type', 'features']

        validator = ExcelValidator(required_columns, optional_columns)
        result = ImportResult()

        try:
            validator.validate_file_format(file)
            headers, rows_data = await self.excel_reader.read_excel_file(file)

            missing_columns = validator.validate_headers(headers)
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required columns: {', '.join(missing_columns)}"
                )

            # Process rows
            for row_number, row_data in enumerate(rows_data, start=2):
                try:
                    clean_data = validator.normalize_row_data(row_data)

                    # Validate required fields
                    if not clean_data.get('room_number'):
                        result.add_error(row_number, "Room number is required", clean_data)
                        continue

                    if not clean_data.get('capacity'):
                        result.add_error(row_number, "Capacity is required", clean_data)
                        continue

                    # Check for duplicate room number
                    existing_room = db.query(Classroom).filter(
                        Classroom.room_number == clean_data['room_number'],
                        Classroom.institution_id == institution_id,
                        Classroom.deleted_at.is_(None)
                    ).first()

                    if existing_room:
                        result.add_error(
                            row_number,
                            f"Room '{clean_data['room_number']}' already exists",
                            clean_data
                        )
                        continue

                    # Prepare room data
                    room_data = {
                        'institution_id': institution_id,
                        'room_number': clean_data['room_number'],
                        'building': clean_data.get('building'),
                        'capacity': int(clean_data['capacity']),
                        'room_type': clean_data.get('room_type', 'lecture_hall'),
                        'features': []
                    }

                    # Process features
                    if clean_data.get('features'):
                        features_str = str(clean_data['features'])
                        room_data['features'] = [
                            f.strip().lower() for f in features_str.split(',') if f.strip()
                        ]

                    # Validate with Pydantic schema
                    room_schema = RoomCreate(**room_data)

                    # Create room record
                    room = Classroom(**room_schema.model_dump())
                    db.add(room)
                    db.flush()

                    result.add_success({
                        'room_number': room.room_number,
                        'building': room.building,
                        'capacity': room.capacity,
                        'id': str(room.id)
                    })

                except ValueError as e:
                    result.add_error(row_number, f"Data validation error: {str(e)}", clean_data)
                except Exception as e:
                    result.add_error(row_number, f"Unexpected error: {str(e)}", clean_data)

            # Commit or rollback
            if result.successful > 0:
                db.commit()
            else:
                db.rollback()

        except HTTPException:
            db.rollback()
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

        return result

    def generate_faculty_template(self) -> io.BytesIO:
        """Generate Excel template for faculty import."""
        headers = [
            'employee_id', 'name', 'email', 'department_code',
            'designation', 'max_hours_per_week', 'subjects_can_teach'
        ]

        sample_data = [
            {
                'employee_id': 'FAC001',
                'name': 'Dr. John Smith',
                'email': 'john.smith@university.edu',
                'department_code': 'CSE',
                'designation': 'Associate Professor',
                'max_hours_per_week': '20',
                'subjects_can_teach': 'Computer Science, Data Structures, Algorithms'
            },
            {
                'employee_id': 'FAC002',
                'name': 'Prof. Sarah Johnson',
                'email': 'sarah.johnson@university.edu',
                'department_code': 'ECE',
                'designation': 'Professor',
                'max_hours_per_week': '18',
                'subjects_can_teach': 'Electronics, Digital Circuits'
            }
        ]

        validations = {
            'designation': [
                'Professor', 'Associate Professor', 'Assistant Professor',
                'Lecturer', 'Senior Lecturer', 'Instructor'
            ]
        }

        instructions = [
            "Fill in all required columns: employee_id, name, department_code",
            "Employee ID must be unique within the institution",
            "Department code must match existing departments",
            "Email addresses should be valid format",
            "Max hours per week should be a number (default: 18)",
            "Subjects can teach: separate multiple subjects with commas",
            "Leave optional fields blank if not applicable"
        ]

        return self.template_generator.create_template(
            headers, sample_data, validations, instructions
        )

    def generate_course_template(self) -> io.BytesIO:
        """Generate Excel template for course import."""
        headers = [
            'code', 'name', 'department_code', 'course_type',
            'theory_credits', 'lab_credits', 'hours_per_week',
            'faculty_employee_id', 'batch_name', 'expected_students', 'required_features'
        ]

        sample_data = [
            {
                'code': 'CS101',
                'name': 'Introduction to Computer Science',
                'department_code': 'CSE',
                'course_type': 'theory',
                'theory_credits': '3.0',
                'lab_credits': '1.0',
                'hours_per_week': '4',
                'faculty_employee_id': 'FAC001',
                'batch_name': 'CSE-2024-A',
                'expected_students': '45',
                'required_features': 'projector, whiteboard, computer'
            }
        ]

        validations = {
            'course_type': ['theory', 'lab', 'tutorial'],
            'required_features': [
                'projector', 'whiteboard', 'computer', 'internet',
                'audio_system', 'microphone', 'air_conditioning'
            ]
        }

        instructions = [
            "Fill in all required columns: code, name, department_code, hours_per_week",
            "Course code must be unique within the institution",
            "Department code must match existing departments",
            "Hours per week should be a positive integer",
            "Faculty employee ID and batch name are optional but must exist if provided",
            "Required features: separate multiple features with commas"
        ]

        return self.template_generator.create_template(
            headers, sample_data, validations, instructions
        )

    def generate_room_template(self) -> io.BytesIO:
        """Generate Excel template for room import."""
        headers = [
            'room_number', 'building', 'capacity', 'room_type', 'features'
        ]

        sample_data = [
            {
                'room_number': 'CS-101',
                'building': 'Computer Science Building',
                'capacity': '50',
                'room_type': 'computer_lab',
                'features': 'projector, computer, internet, air_conditioning'
            },
            {
                'room_number': 'LH-201',
                'building': 'Main Academic Block',
                'capacity': '100',
                'room_type': 'lecture_hall',
                'features': 'projector, whiteboard, audio_system'
            }
        ]

        validations = {
            'room_type': [
                'lecture_hall', 'computer_lab', 'physics_lab',
                'chemistry_lab', 'seminar_room', 'auditorium'
            ],
            'features': [
                'projector', 'whiteboard', 'computer', 'internet',
                'audio_system', 'microphone', 'air_conditioning',
                'laboratory_equipment', 'smart_board'
            ]
        }

        instructions = [
            "Fill in all required columns: room_number, capacity",
            "Room number must be unique within the institution",
            "Capacity should be a positive integer",
            "Room type: select from dropdown or leave blank for 'lecture_hall'",
            "Features: separate multiple features with commas"
        ]

        return self.template_generator.create_template(
            headers, sample_data, validations, instructions
        )

    def export_faculty_excel(
        self,
        db: Session,
        faculty_list: List[Faculty]
    ) -> io.BytesIO:
        """
        Export faculty data to Excel file.

        Args:
            db: Database session
            faculty_list: List of faculty to export

        Returns:
            BytesIO object containing Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Faculty Export"

        # Define headers
        headers = [
            'Employee ID', 'Name', 'Email', 'Department Code',
            'Department Name', 'Designation', 'Max Hours Per Week',
            'Subjects Can Teach', 'Current Workload', 'Available Hours',
            'Utilization %', 'Status'
        ]

        # Style for headers
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        row_idx = 2
        for faculty in faculty_list:
            # Get department info
            department_code = faculty.department.code if faculty.department else "N/A"
            department_name = faculty.department.name if faculty.department else "N/A"

            # Get workload info
            from app.services.faculty_service import faculty_service
            workload_info = faculty_service.get_workload(db, faculty.id)
            current_workload = workload_info["assigned_hours"] if workload_info else 0
            available_hours = workload_info["available_hours"] if workload_info else faculty.max_hours_per_week
            utilization = workload_info["utilization_percentage"] if workload_info else 0.0
            status = "Overloaded" if workload_info and workload_info["overloaded"] else "Available"

            # Format subjects
            subjects = ", ".join(faculty.subjects_can_teach) if faculty.subjects_can_teach else ""

            # Write row data
            row_data = [
                faculty.employee_id,
                faculty.name,
                faculty.email or "",
                department_code,
                department_name,
                faculty.designation or "",
                faculty.max_hours_per_week,
                subjects,
                current_workload,
                available_hours,
                f"{utilization}%",
                status
            ]

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            row_idx += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# Create singleton instance
import_service = ImportService()