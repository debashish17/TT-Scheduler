"""
Excel utility functions for import/export operations.
Provides common functionality for reading Excel files and generating templates.
"""
import io
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import UploadFile, HTTPException


class ExcelValidator:
    """Validates Excel file structure and data."""

    def __init__(self, required_columns: List[str], optional_columns: List[str] = None):
        self.required_columns = required_columns
        self.optional_columns = optional_columns or []
        self.all_columns = required_columns + self.optional_columns

    def validate_file_format(self, file: UploadFile) -> None:
        """Validate file format and size."""
        if not file.filename:
            raise HTTPException(status_code=400, detail="No filename provided")

        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="Only Excel files (.xlsx, .xls) are allowed"
            )

    def validate_headers(self, headers: List[str]) -> List[str]:
        """Validate Excel headers and return missing required columns."""
        missing_columns = []
        normalized_headers = [h.strip().lower() if h else '' for h in headers]

        for required_col in self.required_columns:
            if required_col.lower() not in normalized_headers:
                missing_columns.append(required_col)

        return missing_columns

    def normalize_row_data(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and normalize row data."""
        normalized = {}

        for key, value in row_data.items():
            # Clean key
            clean_key = key.strip().lower().replace(' ', '_') if key else ''

            # Clean value
            if value is None:
                clean_value = None
            elif isinstance(value, str):
                clean_value = value.strip() if value.strip() else None
            else:
                clean_value = value

            normalized[clean_key] = clean_value

        return normalized


class ExcelReader:
    """Reads and parses Excel files safely."""

    @staticmethod
    async def read_excel_file(
        file: UploadFile,
        sheet_name: Optional[str] = None,
        max_rows: int = 10000
    ) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Read Excel file and return headers and row data.

        Args:
            file: Uploaded Excel file
            sheet_name: Specific sheet to read (uses first sheet if None)
            max_rows: Maximum number of rows to process

        Returns:
            Tuple of (headers, rows_data)
        """
        try:
            # Read file content
            contents = await file.read()

            # Load workbook
            workbook = load_workbook(io.BytesIO(contents), read_only=True)

            # Get worksheet
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active

            # Read headers (first row)
            headers = []
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                headers = [str(cell) if cell is not None else '' for cell in first_row]

            # Read data rows
            rows_data = []
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                if row_num >= max_rows:
                    break

                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue

                # Create row dict
                row_dict = {}
                for i, cell_value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell_value

                rows_data.append(row_dict)

            workbook.close()
            return headers, rows_data

        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Error reading Excel file: {str(e)}"
            )


class ExcelTemplateGenerator:
    """Generates Excel templates for data import."""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_template(
        self,
        headers: List[str],
        sample_data: List[Dict[str, Any]] = None,
        validations: Dict[str, List[str]] = None,
        instructions: List[str] = None
    ) -> io.BytesIO:
        """
        Create an Excel template with headers, sample data, and validations.

        Args:
            headers: Column headers
            sample_data: Sample rows for reference
            validations: Data validation rules {column: allowed_values}
            instructions: Instructions for users

        Returns:
            BytesIO object containing the Excel file
        """
        workbook = Workbook()

        # Data sheet
        data_sheet = workbook.active
        data_sheet.title = "Data"

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = data_sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        # Add sample data
        if sample_data:
            for row_num, sample_row in enumerate(sample_data, 2):
                for col_num, header in enumerate(headers, 1):
                    cell = data_sheet.cell(row=row_num, column=col_num)
                    cell.value = sample_row.get(header, "")
                    cell.border = self.border

        # Apply data validations
        if validations:
            self._apply_data_validations(data_sheet, headers, validations)

        # Auto-size columns
        self._auto_size_columns(data_sheet)

        # Add instructions sheet
        if instructions:
            self._add_instructions_sheet(workbook, instructions)

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def _apply_data_validations(
        self,
        sheet,
        headers: List[str],
        validations: Dict[str, List[str]]
    ):
        """Apply data validation rules to columns."""
        for header, allowed_values in validations.items():
            if header in headers:
                col_num = headers.index(header) + 1
                col_letter = get_column_letter(col_num)

                # Create validation rule
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(allowed_values)}"',
                    showDropDown=True
                )
                dv.error = f"Please select from: {', '.join(allowed_values)}"
                dv.errorTitle = "Invalid Value"

                # Apply to column (rows 2-1000)
                dv.add(f"{col_letter}2:{col_letter}1000")
                sheet.add_data_validation(dv)

    def _auto_size_columns(self, sheet):
        """Auto-size columns based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _add_instructions_sheet(self, workbook: Workbook, instructions: List[str]):
        """Add an instructions sheet to the workbook."""
        instructions_sheet = workbook.create_sheet("Instructions")

        # Title
        title_cell = instructions_sheet.cell(row=1, column=1)
        title_cell.value = "Import Instructions"
        title_cell.font = Font(bold=True, size=16)

        # Instructions
        for i, instruction in enumerate(instructions, 3):
            cell = instructions_sheet.cell(row=i, column=1)
            cell.value = f"{i-2}. {instruction}"

        # Auto-size
        self._auto_size_columns(instructions_sheet)


class ImportResult:
    """Represents the result of an import operation."""

    def __init__(self):
        self.total_processed = 0
        self.successful = 0
        self.failed = 0
        self.errors: List[Dict[str, Any]] = []
        self.warnings: List[Dict[str, Any]] = []
        self.created_records: List[Dict[str, Any]] = []

    def add_success(self, record_data: Dict[str, Any]):
        """Add a successful import record."""
        self.successful += 1
        self.total_processed += 1
        self.created_records.append(record_data)

    def add_error(self, row_number: int, error_message: str, row_data: Dict[str, Any] = None):
        """Add an import error."""
        self.failed += 1
        self.total_processed += 1
        self.errors.append({
            "row": row_number,
            "error": error_message,
            "data": row_data or {}
        })

    def add_warning(self, row_number: int, warning_message: str, row_data: Dict[str, Any] = None):
        """Add an import warning."""
        self.warnings.append({
            "row": row_number,
            "warning": warning_message,
            "data": row_data or {}
        })

    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary for API response."""
        return {
            "total_processed": self.total_processed,
            "successful": self.successful,
            "failed": self.failed,
            "errors": self.errors,
            "warnings": self.warnings,
            "success_rate": (self.successful / self.total_processed * 100) if self.total_processed > 0 else 0,
            "created_records": len(self.created_records)
        }